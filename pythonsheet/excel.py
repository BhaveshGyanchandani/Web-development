from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb=Workbook()

print(wb.active.title)
print(wb.sheetnames)
wb['Sheet'].title="new data"   #cganges the sheet name
sh1=wb.active

a1=sh1['A1'].value="name"
print(a1)
sh1['A1'].fill=PatternFill("solid",fgColor="71FF33")
#and etc to creatre teh column in a sheet

# wb.save("data.xlsx")  overwrites the older one

# wb.save("newdat2.xlsx")  