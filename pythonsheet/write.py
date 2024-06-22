import openpyxl

wb=openpyxl.load_workbook("E:\\web development\\pythonsheet\\data.xlsx")#loads the workbook

sh1=wb['data']
rows=sh1.max_row
column=sh1.max_column


sh1.cell(row=2,column=1,value='UK')
sh1.cell(row=2,column=2,value='K')
sh1.cell(row=2,column=3,value='U')

for i in range ( 1 , rows+1):
    for j in range ( 1 , column+1):
        print(sh1.cell(i,j).value)

# wb.save("data.xlsx")  overwrites the older one

# wb.save("newdat.xlsx")  creates the new file and saves old + new data entered
#its better to give the path to save the file