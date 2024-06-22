import openpyxl 

wb=openpyxl.load_workbook("E:\\web development\\pythonsheet\\data.xlsx")#loads the workbook
print(type(wb))
sheets=wb.sheetnames
print(sheets)
print(wb.active.title) #gives the name of current active sheet
sh1=wb['data']  # goimng to  aparticular worksheet
# a1=sh1.cell(1,1).value   give the data of that particular cell
a1=sh1['A1'].value
print(a1)
# print(sh1['A1'].value)

# print(wb['sheert_name']['cell'].value)

print(wb['data']['B1'].value)

print(sh1.cell(row=1,column=1).value)

# print(wb.get_sheet_by_name('data').cell(row=1,column=2).value)    runs but givers an depricated warning